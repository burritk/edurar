import requests

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def create_spreadsheet(category, cycles, output, fcategory=''):
    wb = Workbook()
    ws = wb.active
    row = 1
    ws[get_column_letter(1) + str(row)] = 'Categories'
    ws[get_column_letter(2) + str(row)] = 'Subcategories'
    ws[get_column_letter(3) + str(row)] = 'Company Name'
    ws[get_column_letter(4) + str(row)] = 'Description'
    ws[get_column_letter(5) + str(row)] = 'Website'
    ws[get_column_letter(6) + str(row)] = 'E-mail'
    ws[get_column_letter(7) + str(row)] = 'Phone'
    ws[get_column_letter(8) + str(row)] = 'Address'
    row = 2
    for i in range(0, cycles):
        rows_per_page = 0 + i*5
        page_num = i+1
        r = requests.post('https://www.edurar.com/edirectory/SearchCompanyResult',
                          json={
                              "RowsPerPage": rows_per_page,
                              "PageNo": page_num,
                              "keyword": "",
                              "area": "",
                              "category": category,
                              "prodservice": "",
                              "brand": "",
                              "company": "",
                              "building": "",
                              "advarea": "",
                              "fcategory": fcategory,
                              "farea": "",
                              "frating": "",
                              "fprodservice": "",
                              "fbrand": "",
                              "ftype": "",
                              "searchType": "advanced",
                              "Lat": "",
                              "Logt": ""
                          })

        jsonlist = r.json()
        for dict in jsonlist:
            categories = '\n'.join([item.split('~')[0] for item in dict['ListCategories'].split('|')])
            subcategories = '\n'.join([item.split('~')[0] for item in dict['KeywordProduct'].split('|')])
            company_name = dict['CompanyName']
            description = dict['FullAddress']
            website = dict['Website']
            email = dict['Email']
            phone = dict['Telephone']
            address = dict['AreaName']

            print categories, subcategories, company_name, description, website, email, phone, address

            ws[get_column_letter(1) + str(row)] = categories
            ws[get_column_letter(2) + str(row)] = subcategories
            ws[get_column_letter(3) + str(row)] = company_name
            ws[get_column_letter(4) + str(row)] = description
            ws[get_column_letter(5) + str(row)] = website
            ws[get_column_letter(6) + str(row)] = email
            ws[get_column_letter(7) + str(row)] = phone
            ws[get_column_letter(8) + str(row)] = address
            row += 1

    wb.save(output + '.xlsx')

# create_spreadsheet("920~Specialty Stores", 14, 'specialty_stores')
# create_spreadsheet("2~Manufacturing", 148, 'manufacturing_clothing', fcategory='731')
# create_spreadsheet("2~Manufacturing", 188, 'manufacturing_textile', fcategory='183')
# create_spreadsheet("2~Manufacturing", 12, 'manufacturing_cloths', fcategory='913')
# create_spreadsheet("2~Manufacturing", 27, 'manufacturing_garments', fcategory='191')
# create_spreadsheet("731~Clothing, Shoes, Leather and Bags", 1132, 'clothing_shoes')
create_spreadsheet("847~Restaurants and Coffee Shops", 811, 'restaurants_coffee_shops')
create_spreadsheet("847~Restaurants and Coffee Shops", 116, 'perfumes')
create_spreadsheet("847~Restaurants and Coffee Shops", 196, 'travel_tourism')
create_spreadsheet("847~Restaurants and Coffee Shops", 492, 'cosmetic_beauty')





